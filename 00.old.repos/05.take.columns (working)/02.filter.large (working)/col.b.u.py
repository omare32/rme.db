import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook("0169.culverts.RME_Material_Movement_with_cos_220523.xlsx")

# Get the source sheet
sheet1 = wb.active

# Get the column numbers to copy
columnB = 2
columnU = 21

# Create a new sheet
sheet2 = wb.create_sheet("New Sheet")

# Copy the data from the source sheet to the new sheet
for row in range(1, sheet1.max_row + 1):
    sheet2.cell(row=row, column=1).value = sheet1.cell(row=row, column=columnB).value
    sheet2.cell(row=row, column=2).value = sheet1.cell(row=row, column=columnU).value

# Save the Excel file
wb.save("0169.culverts.RME_Material_Movement_with_cos_220523.new.xlsx")