# %%
# step 01 : from erp (xls/html) to xlsx

import os
import pandas as pd
from bs4 import BeautifulSoup
import xlsxwriter

# Get all .xls files in the current directory
xls_files = [file for file in os.listdir() if file.endswith(".xls")]

# Loop over the .xls files and rename them to .html files
for xls_file in xls_files:
    # Get the new file name
    new_file_name = xls_file[:-4] + ".html"

    # Rename the file
    os.rename(xls_file, new_file_name)
	
# Get all HTML files in the current directory
html_files = [file for file in os.listdir() if file.endswith(".html")]

# Loop over the HTML files
for html_file in html_files:
    # Read the HTML file
    with open(html_file, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f, "html.parser")

    # Get the data in the HTML file
    data = soup.find_all("table")

    # Create a list of lists to store the data
    data_list = []
    for table in data:
        rows = table.find_all("tr")
        for row in rows:
            cols = row.find_all("td")
            data_list.append([col.text for col in cols])

    # Create a Pandas DataFrame from the data
    df = pd.DataFrame(data_list)

    # Save the DataFrame to an Excel file
    df.to_excel(html_file[:-5] + " (python).xlsx", engine="xlsxwriter")
	
# Delete all HTML files in the current directory
for file in os.listdir():
    if file.endswith(".html"):
        os.remove(file)

# %%
# step 02 : remove 1 col and 12 rows

def remove_first_one_column_and_four_rows(file_path):
    """
    Remove the first 1 column and first 12 rows from an Excel file without saving the index row and number index.

    Args:
        file_path (str): The path to the Excel file.
    """

    df = pd.read_excel(file_path)
    df = df.iloc[13:, 1:]
    df = df.reset_index(drop=True)
    df.to_excel(file_path, index=False, header=None)


for file in os.listdir():
    if file.endswith(".xlsx"):
        file_path = os.path.join(os.getcwd(), file)
        remove_first_one_column_and_four_rows(file_path)

# %%
# step 03 : rename headers for all columns

import glob

# Get the list of Excel files in the current directory
excel_files = glob.glob('*.xlsx')  # Update the file extension if needed

# Assuming only one Excel file is present in the directory
if len(excel_files) == 1:
    file_name = excel_files[0]
    
    # Read the Excel file
    data = pd.read_excel(file_name)
    
    # Define the new column names
    new_column_names = [
        'trs_id', 'transaction_source', 'project_no', 'project_name', 'project_zone', 'task_no', 'task_name',
        'top_task_no', 'top_task_name', 'po_no', 'gl_date', 'expenditure_type', 'project_location',
        'project_floor', 'project_area', 'expenditure_category', 'expend_org', 'amount', 'line_no',
        'line_desc', 'inv_no', 'unit', 'qty', 'ipc_no', 'supplier_no', 'supplier_name', 'supplier_site',
        'comment', 'inventory_item', 'owner', 'distributions_status', 'distributions_date', 'distributions_details'
    ]

    # Rename the columns
    data.columns = new_column_names

    # Save the updated DataFrame back to the Excel file
    data.to_excel(file_name, index=False)
    print(f"Column headers in '{file_name}' have been renamed.")
else:
    print("There are zero or multiple Excel files in the directory. Please ensure there is exactly one Excel file.")


# %%
# Step 04 : fix number as text

import openpyxl as px

# Get a list of all Excel files in the current directory
excel_files = [file for file in os.listdir() if file.endswith('.xlsx')]

if not excel_files:
    print("No Excel files found in the current directory.")
else:
    # Select the first Excel file from the list
    file_name = excel_files[0]

    # Load the Excel file
    workbook = px.load_workbook(file_name)
    sheet = workbook.active

    # Iterate through the cells in column R (column 18) and remove commas, convert to numbers (amount)
    for row in sheet.iter_rows(min_row=2, min_col=18, max_col=18):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str):
                cell.value = cell.value.replace(',', '')
                try:
                    cell.value = float(cell.value)
                except ValueError:
                    pass

    # Iterate through the cells in column W (column 23) and remove commas, convert to numbers (qty)
    for row in sheet.iter_rows(min_row=2, min_col=23, max_col=23):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str):
                cell.value = cell.value.replace(',', '')
                try:
                    cell.value = float(cell.value)
                except ValueError:
                    pass

    # Save the modified Excel file with the same name as the old one
    workbook.save(file_name)

    print(f"Fixed Excel file saved as: {file_name}")

# %%
# step 05 : filter on marksup and create tab

from openpyxl import load_workbook

# Find the Excel file in the current directory
excel_files = glob.glob("*.xlsx")

if not excel_files:
    print("No Excel file found in the directory.")
else:
    # Take the first Excel file found (you might adjust this logic if needed)
    file_name = excel_files[0]

    # Read the Excel file
    data = pd.read_excel(file_name)

    # Clean the 'expenditure_type' column by stripping whitespace
    data['expenditure_type'] = data['expenditure_type'].astype(str).str.strip()

    # Define the main criteria for filtering
    criteria = ["Bank Fees", "Stamp Duties", "Insurance", "Fees - stamp"]

    # Filter the data based on the cleaned 'expenditure_type' column
    filtered_data = data[data['expenditure_type'].isin(criteria)]

    # Create a new Excel file and write the filtered data to a new tab called 'marksup'
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
        filtered_data.to_excel(writer, sheet_name='marksup', index=False)

        # Calculate the sum of the 'amount' column for the filtered data
        total_amount = filtered_data['amount'].sum()

        # Write the total amount under the table of the filtered data in the 'marksup' tab
        workbook = writer.book
        worksheet = workbook['marksup']
        max_row = worksheet.max_row

        worksheet.cell(row=max_row + 2, column=1).value = "Total Amount"
        worksheet.cell(row=max_row + 2, column=2).value = total_amount

    # Save the changes using openpyxl
    book = load_workbook(file_name)
    book.save(file_name)




# %%
# step 06 : filter on penalties and create tab

from openpyxl import load_workbook

# Find the Excel file in the current directory
excel_files = glob.glob("*.xlsx")

if not excel_files:
    print("No Excel file found in the directory.")
else:
    # Take the first Excel file found (you might adjust this logic if needed)
    file_name = excel_files[0]

    # Read the Excel file
    data = pd.read_excel(file_name)

    # Clean the 'expenditure_type' column by stripping whitespace
    data['expenditure_type'] = data['expenditure_type'].astype(str).str.strip()

    # Define criteria for penalties filtering
    penalties_criteria = ["Penalty", "Penalty- Utilities", "Penalty- Row Material"]

    # Filter for 'penalties'
    penalties_data = data[data['expenditure_type'].isin(penalties_criteria)]

    # Create a new Excel file and write the penalties data to a new tab called 'penalties'
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
        penalties_data.to_excel(writer, sheet_name='penalties', index=False)

        # Calculate the sum of the 'amount' column for the penalties data
        total_penalties_amount = penalties_data['amount'].sum()

        # Write the total amount under the table of the penalties data in the 'penalties' tab
        workbook = writer.book
        penalties_worksheet = workbook['penalties']
        max_row_penalties = penalties_worksheet.max_row

        penalties_worksheet.cell(row=max_row_penalties + 2, column=1).value = "Total Penalties Amount"
        penalties_worksheet.cell(row=max_row_penalties + 2, column=2).value = total_penalties_amount

    # Save the changes using openpyxl
    book = load_workbook(file_name)
    book.save(file_name)



# %%
# step 07 : create pdf report of amount/marksup/penalties

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

# Find the Excel file in the current directory
current_directory = os.getcwd()
excel_files = [file for file in os.listdir(current_directory) if file.endswith(".xlsx")]

if not excel_files:
    print("No Excel file found in the directory.")
else:
    file_name = excel_files[0]

    # Load the Excel file containing 'marksup' and 'penalties' tabs
    book = load_workbook(file_name)

    # Get the first tab data
    first_tab_df = pd.read_excel(file_name, sheet_name=book.sheetnames[0])

    # Get the project name
    project_name = first_tab_df['project_name'].iloc[0]

    # Get the 'marksup' and 'penalties' data
    marks_df = pd.read_excel(file_name, sheet_name='marksup')
    penalties_df = pd.read_excel(file_name, sheet_name='penalties')

    # Calculate total original amount from the first tab
    original_amount = first_tab_df['amount'].sum()

    # Calculate total amount for 'marksup' and 'penalties'
    total_marksup = marks_df['amount'].sum()
    total_penalties = penalties_df['amount'].sum()

    # Calculate total direct and indirect amount
    total_direct_indirect = original_amount - total_marksup - total_penalties

    # Create a new worksheet for the report
    report_sheet = book.create_sheet('Report')

    # Write the report data to the worksheet with correct formatting and title
    report_sheet.append(['Project Name', project_name])  # Title: Project Name
    report_sheet.append(['Cost Totals'])  # Subtitle: Cost Totals
    report_sheet.append([])  # Empty row for spacing
    report_sheet.append(['Category', 'Total Amount'])
    report_sheet.append(['Original Amount', f'{original_amount:.0f}'])
    report_sheet.append(['Total Marksup', f'{total_marksup:.0f}'])
    report_sheet.append(['Total Penalties', f'{total_penalties:.0f}'])
    report_sheet.append(['Total Direct and Indirect', f'{total_direct_indirect:.0f}'])

    # Apply a table style to the report
    tab = Table(displayName="ReportTable", ref="A4:B8")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    report_sheet.add_table(tab)

    # Save changes to the Excel file
    book.save(file_name)

    # Create a PDF from the report worksheet
    report_pdf_path = 'Report_Total_Amounts.pdf'
    report_canvas = canvas.Canvas(report_pdf_path, pagesize=letter)

    # Get PDF page size
    width, height = letter

    # Define data for the report
    data = [
        ["Project Name =", project_name],
        ["Cost Totals"],
        [],
        ["Original Amount =", f'{original_amount:,.0f}'],
        ["Total Marksup =", f'{total_marksup:,.0f}'],
        ["Total Penalties =", f'{total_penalties:,.0f}'],
        ["Total Direct and Indirect =", f'{total_direct_indirect:,.0f}']
    ]

    # Set positions for text elements
    text_start_y = height - 50  # Start the text from 50 pixels from the top
    text_gap = 20

    # Draw text elements on the PDF
    text = report_canvas.beginText(0, text_start_y)
    text.setFont("Helvetica-Bold", 12)  # Title: Bold and larger font size
    max_text_width = 0
    for row in data:
        text_width = report_canvas.stringWidth(" ".join(map(str, row)), "Helvetica-Bold", 12)
        if text_width > max_text_width:
            max_text_width = text_width

    text_start_x = (width - max_text_width) / 2  # Center align horizontally
    text.setTextOrigin(text_start_x, text_start_y)

    for i, row in enumerate(data):
        if i == 0:
            text.setFont("Helvetica-Bold", 12)  # Title: Bold and larger font size
        else:
            text.setFont("Helvetica", 10)  # Values: Smaller font size
        text.textLine(" ".join(map(str, row)))
        text_start_y -= text_gap

    report_canvas.drawText(text)
    report_canvas.save()




# %%
# step 8 : steel rft qty and amount

import os
import openpyxl

# Get the current directory
current_directory = os.getcwd()

# Find the Excel file in the current directory
excel_file = None
for file in os.listdir(current_directory):
    if file.endswith('.xlsx'):
        excel_file = file
        break

if excel_file:
    # Load the Excel workbook
    wb = openpyxl.load_workbook(excel_file)

    # Rename the "Sheet1" to "cost_dist_all"
    if "Sheet1" in wb.sheetnames:
        ws = wb["Sheet1"]
        ws.title = "cost_dist_all"

    # Create a new tab called "steel rft"
    wb.create_sheet(title="steel rft")

    # Save the changes to the workbook
    wb.save(excel_file)
    print("Excel file updated successfully.")
else:
    print("No Excel file found in the current directory.")    

# %%
# step 8 part 2 (8b)

import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment

# Get the current directory
current_directory = os.getcwd()

# Find the Excel file in the current directory
excel_file = None
for file in os.listdir(current_directory):
    if file.endswith('.xlsx'):
        excel_file = file
        break

if excel_file:
    # Load the Excel workbook
    wb = openpyxl.load_workbook(excel_file)

    # Reference to sheets
    cost_dist_all = wb["cost_dist_all"]

    # Check if "comment" column exists in cost_dist_all sheet
    comment_column_index = None
    for col in cost_dist_all.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if "comment" in str(cell.value).lower():
                comment_column_index = cell.column
                break
        if comment_column_index:
            break

    if comment_column_index:
        # Find indices for required columns
        column_indices = {}
        headers = [cell.value.lower() for cell in cost_dist_all[1]]
        for col_name in ["gl_date", "comment", "unit", "qty", "amount"]:
            if col_name in headers:
                column_indices[col_name] = headers.index(col_name) + 1

        # Check if "steel rft" sheet already exists, if yes, delete it
        if "steel rft" in wb.sheetnames:
            wb.remove(wb["steel rft"])

        # Create a new "steel rft" sheet
        steel_rft = wb.create_sheet(title="steel rft")

        # Create headers for steel_rft sheet
        steel_rft.append(["gl_date", "comment", "unit", "rate", "qty", "amount"])

        # Iterate through rows and copy relevant data to steel_rft sheet
        for row in cost_dist_all.iter_rows(min_row=2, values_only=True):
            if "حديد تسليح" in str(row[comment_column_index - 1]).lower():
                gl_date = row[column_indices.get("gl_date", 0) - 1]
                comment = row[column_indices.get("comment", 0) - 1]
                unit = row[column_indices.get("unit", 0) - 1]
                qty = row[column_indices.get("qty", 0) - 1]
                amount = row[column_indices.get("amount", 0) - 1]

                # Calculate rate
                rate = amount / qty if qty != 0 else 0

                # Append the data to the steel_rft sheet
                steel_rft.append([gl_date, comment, unit, rate, qty, amount])

        # Calculate totals and average rate
        total_amount = sum(steel_rft.cell(row=i, column=6).value for i in range(2, steel_rft.max_row + 1))
        total_qty = sum(steel_rft.cell(row=i, column=5).value for i in range(2, steel_rft.max_row + 1))
        average_rate = total_amount / total_qty if total_qty != 0 else 0

        # Add row with totals and average rate
        total_row = ["Total", "", "", average_rate, total_qty, total_amount]
        steel_rft.append(total_row)

        # Apply formatting to the total row
        for cell in steel_rft[steel_rft.max_row]:
            cell.font = Font(size=14, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Formatting for the header row
        for cell in steel_rft[1]:
            cell.font = Font(size=14, bold=True)
            cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Freeze the top row
        steel_rft.freeze_panes = "A2"

        # Set column widths
        steel_rft.column_dimensions["A"].width = 15  # Width for the gl_date column
        steel_rft.column_dimensions["B"].width = 30  # Width for the comment column
        steel_rft.column_dimensions["F"].width = 25  # Width for the amount column

        # Save the changes to the workbook
        wb.save(excel_file)
        print("Data consolidated in 'steel rft' sheet with formatting, frozen row, and adjusted column widths.")
    else:
        print("No 'comment' column found in the 'cost_dist_all' sheet.")
else:
    print("No Excel file found in the current directory.")


# %%
# step 8 part 3 (8c)

from openpyxl import load_workbook
from decimal import Decimal
import os
from weasyprint import HTML

excel_file = None
for file in os.listdir():
    if file.endswith('.xlsx'):
        excel_file = file
        break

if excel_file:
    wb = load_workbook(excel_file)
    cost_dist_all = wb['cost_dist_all']

    # Extract project name from the "project_name" column in "cost_dist_all"
    project_name = cost_dist_all['D2'].value  # Assuming project name is in cell D2

    # Replace any problematic characters in the project name
    project_name = project_name.replace('\n', '').replace('/', '').replace(':', '')

    # Get the "steel rft" sheet
    steel_rft = wb['steel rft']  # Assuming sheet name is "steel rft"

    # Extract data from "steel rft" sheet and format numbers
    data = []
    for row in steel_rft.iter_rows(values_only=True):
        formatted_row = []
        for cell in row:
            if isinstance(cell, float):
                formatted_row.append(format(Decimal(cell), ',.0f'))  # Format numbers without decimal places and with commas
            else:
                formatted_row.append(cell)
        data.append(formatted_row)

    # PDF title
    pdf_title = f"<h1>{project_name} Steel Rft Log</h1>"

    # Generate HTML table
    html_table = "<table border='1' cellspacing='0' cellpadding='2' style='text-align:center;'>"
    for i, row in enumerate(data):
        if i == 0:
            html_table += "<thead><tr>"
            for cell in row:
                html_table += f"<th>{cell}</th>"
            html_table += "</tr></thead><tbody>"
        else:
            html_table += "<tr>"
            for cell in row:
                html_table += f"<td>{cell}</td>"
            html_table += "</tr>"
    html_table += "</tbody></table>"

    # Adding space above the last row (total row)
    html_table += "<div style='height: 20px'></div>"

    # Create PDF using WeasyPrint
    pdf_filename = f"{project_name} Steel Rft Log.pdf"
    final_html = f"{pdf_title}{html_table}"  # Combine title and table
    HTML(string=final_html).write_pdf(pdf_filename)

    print(f"PDF '{pdf_filename}' created successfully using WeasyPrint.")



