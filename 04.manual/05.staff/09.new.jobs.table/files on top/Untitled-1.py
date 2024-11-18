# %%
# step 01 : from erp (xls/html) to xlsx

import os
import pandas as pd
from bs4 import BeautifulSoup
import xlsxwriter

# Get all .xls files in the current directory
xls_files = [file for file in os.listdir() if file.endswith(".xls")]

# Loop over the .xls files and rename them to .html files
for xls_file in xls_files:
    # Get the new file name
    new_file_name = xls_file[:-4] + ".html"

    # Rename the file
    os.rename(xls_file, new_file_name)
	
# Get all HTML files in the current directory
html_files = [file for file in os.listdir() if file.endswith(".html")]

# Loop over the HTML files
for html_file in html_files:
    # Read the HTML file
    with open(html_file, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f, "html.parser")

    # Get the data in the HTML file
    data = soup.find_all("table")

    # Create a list of lists to store the data
    data_list = []
    for table in data:
        rows = table.find_all("tr")
        for row in rows:
            cols = row.find_all("td")
            data_list.append([col.text for col in cols])

    # Create a Pandas DataFrame from the data
    df = pd.DataFrame(data_list)

    # Save the DataFrame to an Excel file
    df.to_excel(html_file[:-5] + ".xlsx", engine="xlsxwriter")
	
# Delete all HTML files in the current directory
for file in os.listdir():
    if file.endswith(".html"):
        os.remove(file)

# %%
# step 02 : remove 1 col and 4 rows

def remove_first_one_column_and_four_rows(file_path):
    """
    Remove the first 1 column and first 4 rows from an Excel file without saving the index row and number index.

    Args:
        file_path (str): The path to the Excel file.
    """

    df = pd.read_excel(file_path)
    df = df.iloc[5:, 1:]
    df = df.reset_index(drop=True)
    df.to_excel(file_path, index=False, header=None)


for file in os.listdir():
    if file.endswith(".xlsx"):
        file_path = os.path.join(os.getcwd(), file)
        remove_first_one_column_and_four_rows(file_path)

# %%
# Step 03 : fix number as text

import openpyxl as px

# Get a list of all Excel files in the current directory
excel_files = [file for file in os.listdir() if file.endswith('.xlsx')]

if not excel_files:
    print("No Excel files found in the current directory.")
else:
    # Select the first Excel file from the list
    file_name = excel_files[0]

    # Load the Excel file
    workbook = px.load_workbook(file_name)
    sheet = workbook.active

    # Iterate through the cells in column S (column 19) and remove commas, convert to numbers
    for row in sheet.iter_rows(min_row=2, min_col=19, max_col=19):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str):
                cell.value = cell.value.replace(',', '')
                try:
                    cell.value = float(cell.value)
                except ValueError:
                    pass

    # Iterate through the cells in column U (column 21) and remove commas, convert to numbers
    for row in sheet.iter_rows(min_row=2, min_col=21, max_col=21):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str):
                cell.value = cell.value.replace(',', '')
                try:
                    cell.value = float(cell.value)
                except ValueError:
                    pass

    # Save the modified Excel file with the same name as the old one
    workbook.save(file_name)

    print(f"Fixed Excel file saved as: {file_name}")

# %%
#step 03b : save file name

# Get a list of all files in the current directory
files = os.listdir()

# Filter only the Excel files (assuming they have the '.xlsx' extension)
excel_files = [file for file in files if file.endswith('.xlsx')]

# Check if there is exactly one Excel file
if len(excel_files) == 1:
    excel_file_name00 = excel_files[0]
    print(f"The name of the Excel file is: {excel_file_name00}")
else:
    print("There is not exactly one Excel file in the directory.")

# %%
# step 4 : filter on project in new table

# Get a list of all Excel files in the current directory
excel_files = [file for file in os.listdir() if file.endswith('.xlsx')]

if excel_files:
    # Read the first Excel file found in the directory
    file_to_read = excel_files[0]
    df = pd.read_excel(file_to_read)

    # Clean up column names by removing leading/trailing whitespace
    df.columns = df.columns.str.strip()

    # Define the filter values
    filter_values = ["Move Order Issue on Project", "RME Issue ( On Project)", "RME Site Return"]

    # Filter the DataFrame based on the 'Trx Type' column
    filtered_data = df[df['Trx Type'].str.strip().isin(filter_values)]

    if not filtered_data.empty:
        # Save the filtered data to a new Excel file
        filtered_data.to_excel('filtered_data.xlsx', index=False)
        print("Filtered data saved to 'filtered_data.xlsx'")
    else:
        print("No rows match the filter criteria.")
else:
    print("No Excel files found in the current directory.")

# %%
# step 05 : fix files into one with tabs

# Get a list of all Excel files in the current directory
excel_files = [file for file in os.listdir() if file.endswith('.xlsx')]

# Ensure there are exactly two Excel files in the directory
if len(excel_files) != 2:
    print("There should be exactly two Excel files in the directory.")
else:
    # Sort the files alphabetically to identify them
    excel_files.sort()

    # Define the new Excel file name (change this as needed)
    new_excel_name = "combined_data.xlsx"

    # Create a new Excel writer
    with pd.ExcelWriter(new_excel_name, engine='openpyxl') as writer:
        # Read the first Excel file (filtered_data.xlsx)
        first_excel_data = pd.read_excel(excel_files[0])
        # Write the first data to the first tab ("mat mov")
        first_excel_data.to_excel(writer, sheet_name="mat mov", index=False)

        # Read the second Excel file (the one with the variable name)
        second_excel_data = pd.read_excel(excel_files[1])
        # Write the second data to the second tab ("on project")
        second_excel_data.to_excel(writer, sheet_name="on project", index=False)

    # Delete the original Excel files
    for file in excel_files:
        os.remove(file)

    print(f"The new Excel file '{new_excel_name}' has been created with the specified tabs.")

# %%
# step 05b : fix names (to original name)

# Get a list of all files in the current directory
files = os.listdir()

# Filter only the Excel files (assuming they have the '.xlsx' extension)
excel_files = [file for file in files if file.endswith('.xlsx')]

# Check if there is exactly one Excel file
if len(excel_files) == 1:
    excel_file_name = excel_files[0]
    print(f"The name of the Excel file is: {excel_file_name}")

    # Rename the Excel file to the name in excel_file_name
    new_excel_name = excel_file_name00  # Use the same name
    os.rename(excel_file_name, new_excel_name)

    print(f"The Excel file has been renamed to: {new_excel_name}")
else:
    print("There is not exactly one Excel file in the directory.")

# %%
# step 06 : remove duplicates and sumif

from openpyxl import load_workbook

# Get the current working directory
current_directory = os.getcwd()

# Find the Excel file in the current directory
excel_files = [file for file in os.listdir(current_directory) if file.endswith('.xlsx')]

if len(excel_files) == 0:
    print("No Excel file found in the current directory.")
else:
    # Assuming the first Excel file found is the one to be processed
    excel_file_path = os.path.join(current_directory, excel_files[0])

    # Read Excel file
    xls = pd.ExcelFile(excel_file_path)

    # Read the "on project" tab
    on_project_df = pd.read_excel(xls, 'on project')

    # Create a DataFrame for the "sumif" tab
    sumif_df = pd.DataFrame(columns=['Item Desc', 'Unit', 'Rate', 'Qty', 'Amount'])

    # Get unique values from "Item Desc" column
    unique_items = on_project_df['Item Desc'].unique()

    # Populate "Item Desc" column in the "sumif" tab
    sumif_df['Item Desc'] = unique_items

    # Calculate sum and average for each unique Item Desc
    for item in unique_items:
        item_data = on_project_df[on_project_df['Item Desc'] == item]
        qty_sum = item_data['Quantities'].sum()
        amount_sum = item_data['Total Amount'].sum()
        avg_rate = amount_sum / qty_sum if qty_sum != 0 else 0
        
        # Update values in the "sumif" tab
        sumif_df.loc[sumif_df['Item Desc'] == item, 'Rate'] = avg_rate
        sumif_df.loc[sumif_df['Item Desc'] == item, 'Qty'] = qty_sum
        sumif_df.loc[sumif_df['Item Desc'] == item, 'Amount'] = amount_sum

    # Reorder columns
    sumif_df = sumif_df[['Item Desc', 'Unit', 'Rate', 'Qty', 'Amount']]

    # Write the result to a new tab called "sumif"
    with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a') as writer:
        sumif_df.to_excel(writer, sheet_name='sumif', index=False)

    # Open the workbook again to add the VLOOKUP formula to the "Unit" column
    book = load_workbook(excel_file_path)
    sumif_sheet = book['sumif']

    # Add VLOOKUP formula to the "Unit" column starting from the second row (row_index=2)
    for row_index, item in enumerate(sumif_sheet.iter_rows(min_row=2, max_row=sumif_sheet.max_row, min_col=2, max_col=2), start=2):
        item_desc = sumif_sheet.cell(row=row_index, column=1).value
        formula = f'=VLOOKUP(A{row_index},\'on project\'!B:C,2,0)'
        sumif_sheet.cell(row=row_index, column=2, value=formula)

    # Save the workbook
    book.save(excel_file_path)


# %%
# step 07 : create minor tab and arrange

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

# Get the current working directory
current_directory = os.getcwd()

# List all files in the current directory
files = os.listdir(current_directory)

# Find the first file with a .xlsx extension (assuming there's only one Excel file)
excel_file = next(file for file in files if file.endswith('.xlsx'))

# Construct the full path to the Excel file
file_path = os.path.join(current_directory, excel_file)

# Read the "sumif" sheet into a pandas DataFrame
df_sumif = pd.read_excel(file_path, sheet_name='sumif')

# Select the relevant columns from "sumif" and rename as needed
df_minor = df_sumif[['Item Desc', 'Rate', 'Qty']].copy()
df_minor.rename(columns={'Item Desc': 'Item'}, inplace=True)

# Copy negative values of 'Qty' to the 4th column in the new "minor" sheet
df_minor['Qty'] = -df_minor['Qty']

# Open the workbook and get the "sumif" sheet to access calculated values
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
    workbook = writer.book
    sumif_sheet = workbook['sumif']

    # Extract the "Unit" values from the "sumif" sheet
    unit_values = [cell.value for cell in sumif_sheet['B']]  # Assuming "Unit" is in column B, adjust as needed

    # Create a new sheet named "sorted_minor"
    sorted_minor_sheet = workbook.create_sheet('sorted_minor')

    # Write headers to the "sorted_minor" sheet
    sorted_minor_sheet.append(['Item', 'Unit', 'Rate', 'Qty', 'Amount'])

    # Sort the DataFrame by the "Amount" column from largest to smallest
    df_minor['Amount'] = df_minor['Rate'] * df_minor['Qty']
    df_minor.sort_values(by='Amount', ascending=False, inplace=True)

    # Write data to the "sorted_minor" sheet with correct Excel formulas
    for index, row in enumerate(df_minor.itertuples(), start=2):
        sorted_minor_sheet.append([row.Item, '', row.Rate, row.Qty, f'=C{index}*D{index}'])

    # Write the "Unit" values to the "sorted_minor" sheet
    for index, value in enumerate(unit_values, start=2):
        sorted_minor_sheet[f'B{index}'] = value

    # Apply correct formulas to the "Amount" column in the "sorted_minor" sheet
    for row in sorted_minor_sheet.iter_rows(min_row=2, max_row=sorted_minor_sheet.max_row, min_col=5, max_col=5):
        formula = f'=C{row[0].row}*D{row[0].row}'
        row[0].value = formula
        row[0].alignment = Alignment(horizontal='left')  # Align the formula to the left

print("Data from 'sumif' sheet processed and added to 'sorted_minor' sheet successfully.")

# %%
# step 08 : major and formating

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Get the list of files in the current directory
current_directory = './'
files_in_directory = os.listdir(current_directory)

# Filter Excel files
excel_files = [file for file in files_in_directory if file.endswith('.xlsx')]

if not excel_files:
    raise FileNotFoundError("No Excel files found in the directory.")

# Choose the first Excel file found
file_name = excel_files[0]

# Read the Excel file
file_path = os.path.join(current_directory, file_name)
excel_data = pd.ExcelFile(file_path)

# Check sheet names and rename 'sorted_minor' to 'minor'
sheet_names = excel_data.sheet_names
if 'sorted_minor' in sheet_names:
    df_minor = excel_data.parse('sorted_minor')
    df_minor.rename(columns={'Amount': 'Amount'}, inplace=True)
    df_minor.columns = ['Item', 'Unit', 'Rate', 'Qty', 'Amount']

    # Calculate the product of 'Rate' and 'Qty'
    df_minor['Product'] = df_minor['Rate'] * df_minor['Qty']

    # Filter data for 'major' and 'minor' tabs based on the product
    df_major = df_minor[df_minor['Product'] > 100000]
    df_minor = df_minor[df_minor['Product'] <= 100000]

    # Sort 'major' tab by 'Item'
    df_major = df_major.sort_values('Item')

    # Remove the 'Amount' column from 'major' as 'Product' will be renamed
    if 'Amount' in df_major.columns:
        del df_major['Amount']

    # Rename 'Product' column to 'Amount' in 'major'
    df_major.rename(columns={'Product': 'Amount'}, inplace=True)

    # Create a new ExcelWriter to write the 'minor' and 'major' data
    with pd.ExcelWriter('temp_file.xlsx', engine='openpyxl') as writer:
        df_minor.to_excel(writer, sheet_name='minor', index=False)
        df_major.to_excel(writer, sheet_name='major', index=False)

    # Load the original workbook
    original_wb = load_workbook(file_path)
    temp_wb = load_workbook('temp_file.xlsx')

    # Rename 'sorted_minor' to 'minor'
    if 'sorted_minor' in original_wb.sheetnames:
        original_wb['sorted_minor'].title = 'minor'

    # Copy the 'major' sheet from the temporary workbook to the original workbook
    sheet_source_major = temp_wb['major']
    sheet_destination_major = original_wb.create_sheet('major')

    for row in dataframe_to_rows(df_major, index=False, header=True):
        sheet_destination_major.append(row)

    # Save the changes to the original Excel file
    original_wb.save(file_path)

    # Remove the temporary file
    os.remove('temp_file.xlsx')
else:
    raise ValueError("No 'sorted_minor' sheet found in the Excel file.")




# %%
# step 08b : fix minor and major

import os
import pandas as pd
import shutil

# Get the current working directory
current_directory = os.getcwd()

# Find the Excel file in the current directory
excel_files = [file for file in os.listdir(current_directory) if file.endswith('.xlsx')]

if len(excel_files) == 0:
    print("No Excel file found in the current directory.")
else:
    # Choose the first Excel file found
    file_name = excel_files[0]
    new_file_name = 'updated_' + file_name  # New file name

    # Load the Excel file in read-only mode
    with pd.ExcelFile(file_name) as excel_data:
        # Read the major tab to get the number of rows (excluding header)
        major_tab = excel_data.parse('major')  # Replace 'major' with your major tab name
        num_rows_major = len(major_tab)  # Count includes the header

        # Read the minor tab
        minor_tab = excel_data.parse('minor')  # Replace 'minor' with your minor tab name

        # Delete the extra rows from the minor tab
        minor_tab = minor_tab.iloc[num_rows_major:]  # Slice to remove rows

        # Recalculate 'Amount' column based on 'Qty' and 'Rate'
        minor_tab['Amount'] = minor_tab.iloc[:, 2] * minor_tab.iloc[:, 3]

        # Write the modified data to a new Excel file
        with pd.ExcelWriter(new_file_name, engine='openpyxl') as writer:
            for sheet in excel_data.sheet_names:
                if sheet == 'minor':  # If it's the 'minor' sheet, write the modified data
                    minor_tab.to_excel(writer, sheet_name='minor', index=False)
                else:  # Copy other sheets as they are
                    data = excel_data.parse(sheet)
                    data.to_excel(writer, sheet_name=sheet, index=False)

    # Remove the original file
    os.remove(file_name)

    # Rename the new file to have the original file name
    os.rename(new_file_name, file_name)







