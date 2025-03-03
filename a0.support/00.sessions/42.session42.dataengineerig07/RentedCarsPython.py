from datetime import datetime,timedelta
from azure.storage.blob import BlobServiceClient, generate_blob_sas, BlobSasPermissions,ContainerClient
import pandas as pd
from openpyxl import load_workbook
import openpyxl
import io

#Azure Credentials
account_name="rowad0epsilon"
account_key="iXocBRG8KMTvLHliQZb+GScWzjZ2v+Qx6mIEa/ex5//dgFP/7I0OYiJINDIe2By8xykVc6JL6HaM+AStB2gQ7w=="
container_name="usecase-container"
connect_str = 'DefaultEndpointsProtocol=https;AccountName=' + account_name + ';AccountKey=' + account_key + ';EndpointSuffix=core.windows.net'

#Extract
##############################
def ReadSourceExcelSheetToDataFrame(excelFile,columnNames):

    container_client = ContainerClient.from_connection_string(conn_str=connect_str,container_name=container_name)
    downloaded_blob = container_client.download_blob(excelFile)
    file = io.BytesIO(downloaded_blob.readall())
    wb = load_workbook(file,data_only=True)
    ws = wb[wb.sheetnames[0]]
    print("Source sheet : "+wb.sheetnames[0])
    df = pd.DataFrame(ws.values)
    df = df[2:]
    df = df[:-1]
    df.columns = columnNames
    df["Month"] = datetime.strptime(GetDate(excelFile), "%b-%y").strftime("%Y-%m-%d %H:%M:%S")
    return df

def ReadDestinationExcelSheetToDataFrame(excelFile,sheetName):
    container_client = ContainerClient.from_connection_string(conn_str=connect_str,container_name=container_name)
    downloaded_blob = container_client.download_blob(excelFile)
    file = io.BytesIO(downloaded_blob.readall())
    wb = load_workbook(file)
    ws = wb[sheetName]
    print("Destination sheet : "+sheetName)
    df = pd.DataFrame(ws.values)
    columnNames = df.iloc[0]
    df = df[1:]
    df.columns = columnNames
    return df
##############################

#Transform
##############################
def AppendNewData(SourceDF,DestinationDF,commonColumns):
    commonSourceDF = SourceDF[commonColumns]
    commonDestinationDF = DestinationDF[commonColumns]
    
    #Keep only new Records
    commonSourceDF['Month'] = pd.to_datetime(commonSourceDF['Month'])
    commonDestinationDF['Month'] = pd.to_datetime(commonDestinationDF['Month'])
    
    commonSourceDF = commonSourceDF[commonSourceDF["Month"] > commonDestinationDF["Month"].agg('max')]
    
    if len(commonSourceDF) == 0:
        return pd.DataFrame()
    
    else:
        result = pd.concat([commonDestinationDF,commonSourceDF],axis=0,ignore_index=True)
        # return result
        # Calculated Columns
        result["Cost"] = pd.to_numeric(result["Cost"],errors="coerce")
        result["Cost"] = result["Cost"].astype(float)
        result["Vat 14%"] = result["Cost"] * 0.14
        result["Cost + Vat"] = result["Cost"] + result["Vat 14%"]
        result["Deductions"] = result["Cost"] * 0.03
        result["Cost after Deductions"] = result["Cost + Vat"] - result["Deductions"]
        result["Corona discount"] = 0
        result["Cost After Discount"] = result["Cost after Deductions"]
        result["Fuel"] = 0
        result["Tolls"] = 0
        result["Maintainance"] = 0
        result["Wash"] = 0
        result["Parking"] = 0
        result["Others"] = 0
        result["Running Cost"] = 0
        result["Total Cost"] = result["Cost After Discount"] + result["Running Cost"]
        result["VOW"] = ""
        result["Value Of Work"]= ""
        


        # Formulas
        for (x,y) in zip(range(0,len(result)),range(1,len(result)+1)):
            result.loc[x,"No. Of Seats"]= f"=VLOOKUP($D{y+1},Month!$D:$E,2,0)"
            result.loc[x,"Project Name"]= f"=INDEX('Master Data Sheet'!$A:$H,MATCH('Rented Cars'!$A{y+1},'Master Data Sheet'!$B:$B,0),1)"
            result.loc[x,"P/D/O"] = f"=VLOOKUP($AC{y+1},'Master Data Sheet'!$A:$K,11,0)"
            result.loc[x,"Projects Manager"] = f"=VLOOKUP($AC{y+1},'Master Data Sheet'!$A:$K,9,0)"
            result.loc[x,"UOM"] = f"=VLOOKUP($F{y+1},Month!$G:$H,2,0)"
            result.loc[x,"Location "] = f"=VLOOKUP($AC{y+1},'Master Data Sheet'!$A:$K,10,0)"
            result.loc[x,"Serial"] = f"=VLOOKUP($AC{y+1},'Master Data Sheet'!$A:$H,8,0)"
            
        
        return result
##############################

#Load
##############################
def SaveDataFrameToExcelSheet(DataFrame,excelFile,sheetName,finalColumns):
    if len(DataFrame) == 0:
        return "No New Data to be added"
    
    else:
        DataFrame = DataFrame[finalColumns]
        container_client = ContainerClient.from_connection_string(conn_str=connect_str,container_name=container_name)
        downloaded_blob = container_client.download_blob(excelFile)
        file = io.BytesIO(downloaded_blob.readall())
        with pd.ExcelWriter(file, engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:
            DataFrame.to_excel(writer,sheet_name=sheetName,index=False)

        
        #Create Excel Table
        tab = openpyxl.worksheet.table.Table(displayName="RentedCarsTable", ref=f'A1:{openpyxl.utils.get_column_letter(DataFrame.shape[1])}{len(DataFrame)+1}')
        wb = load_workbook(file)
        wb[sheetName].add_table(tab)
        wb.save(filename=excelFile)


        blob_client = container_client.get_blob_client(blob=excelFile)
        with open(excelFile,"rb") as data:
            blob_client.upload_blob(data,overwrite=True)
        
        return "Added New Records"
##############################

#Helping Functions
##############################
def GetDate(fileName):
    words = fileName.split(' ')
    date= words[-1]
    splitDate = date.split('.')
    finalDate = splitDate[0]+'-'+splitDate[1][2]+splitDate[1][3]
    return finalDate
##############################

sourceExcelFile = "12 Rented cars cost Dec.2023.xlsx"
destinationExcelFile="Transportation & Equipment 2023.xlsx"
sheetName = "Rented Cars"
englishSourceColumns = ["Serial","OLD Project Name","Department/Site","Subcontractor","Car Type","Emp. Name","Unit","Route","Single Rate","Double Rate","Single Qty","Double Qty","Cost","Vat 14%","Deductions 3%","Deductions","Cost after Deductions","Comments"]
commonColumns = ["OLD Project Name","Department/Site","Subcontractor","Car Type","Emp. Name","Unit","Route","Single Rate","Double Rate","Single Qty","Double Qty","Cost","Month"]
finalColumns = ["OLD Project Name","Department/Site","Subcontractor","Car Type","Emp. Name","Unit","Route","Single Rate","Double Rate","Single Qty","Double Qty","Cost","Vat 14%","Cost + Vat","Deductions","Cost after Deductions","Corona discount","Cost After Discount","Fuel","Tolls","Maintainance","Wash","Parking","Others","Running Cost","Total Cost","No. Of Seats","Month","Project Name","P/D/O","Projects Manager","UOM","VOW","Value Of Work","Location ","Serial"]

source = ReadSourceExcelSheetToDataFrame(excelFile=sourceExcelFile,columnNames=englishSourceColumns)
destination = ReadDestinationExcelSheetToDataFrame(excelFile=destinationExcelFile,sheetName=sheetName)
result = AppendNewData(SourceDF=source,DestinationDF=destination,commonColumns=commonColumns)
print(SaveDataFrameToExcelSheet(DataFrame=result,excelFile=destinationExcelFile,sheetName=sheetName,finalColumns=finalColumns))

