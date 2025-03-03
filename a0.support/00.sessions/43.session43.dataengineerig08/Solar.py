from datetime import datetime,timedelta
from azure.storage.blob import BlobServiceClient, generate_blob_sas, BlobSasPermissions,ContainerClient
import pandas as pd
from openpyxl import load_workbook
import openpyxl
import io

#Azure Credentials
account_name="rowad0epsilon"
account_key="iXocBRG8KMTvLHliQZb+GScWzjZ2v+Qx6mIEa/ex5//dgFP/7I0OYiJINDIe2By8xykVc6JL6HaM+AStB2gQ7w=="
container_name="usecase-container"
connect_str = 'DefaultEndpointsProtocol=https;AccountName=' + account_name + ';AccountKey=' + account_key + ';EndpointSuffix=core.windows.net'

#Extract
##############################
def ReadSourceExcelSheetToDataFrame(excelFile):

    container_client = ContainerClient.from_connection_string(conn_str=connect_str,container_name=container_name)
    downloaded_blob = container_client.download_blob(excelFile)
    file = io.BytesIO(downloaded_blob.readall())

    
    df = first_row =  pd.read_excel(file, nrows=20,header=None,index_col=False)
    first_row =  df.iloc[18:19]
    first_row_names= []
    current_column = ""

    for i in range(len(first_row.iloc[0])):
        if pd.notna(first_row.iloc[0][i]):
            current_column = remove(first_row.iloc[0][i])
            first_row_names.append(current_column)
        else:
            first_row_names.append(current_column)



    # Read the first two rows of the CSV file
    first_two_rows = df.iloc[18:20]

    # Combine values from the two rows to create column names
    column_names = []
    for col1, col2 in zip(first_row_names, first_two_rows.iloc[1]):
        if pd.notna(col1) and pd.notna(col2):
            column_names.append(f"{col1}/{col2}")
        elif pd.notna(col1):
            column_names.append(col1)
        elif pd.notna(col2):
            column_names.append(col2)
        

    # Read the excel file skipping the first 20 rows and using the generated column names
    df = pd.read_excel(file, skiprows=20,names=column_names,index_col=False)
    df = df.iloc[:-1]
    
    commonSourceDF_Logistics = df[df['Owner'] == "Logistics  "]
    commonSourceDF_Solar = df[df['LineDesc'] == "سولار"]
    df = pd.concat([commonSourceDF_Logistics,commonSourceDF_Solar],axis=0,ignore_index=True)
    df["GLdate"] = pd.to_datetime(df['GLdate'],unit='d',origin='1899-12-30')
    df["DistributionsDetails/Date"] = pd.to_datetime(df['DistributionsDetails/Date'],unit='d',origin='1899-12-30')
    return df,column_names

def ReadDestinationExcelSheetToDataFrame(excelFile,sheetName,columnNames):
    column_names=columnNames
    container_client = ContainerClient.from_connection_string(conn_str=connect_str,container_name=container_name)
    downloaded_blob = container_client.download_blob(excelFile)
    file = io.BytesIO(downloaded_blob.readall())
    wb = load_workbook(file)
    ws = wb[sheetName]
    print("Destination sheet : "+sheetName)
    df = pd.DataFrame(ws.values)
    df = df[1:]
    column_names.extend(['ProjectName','Serial','ProjectsManager','P/D/O','Location ','Month'])
    df.columns = column_names
    return df,column_names

##############################

#Transform
##############################
def AppendNewData(SourceDF,DestinationDF,commonColumns):

    commonSourceDF = SourceDF[commonColumns]
    commonDestinationDF = DestinationDF[commonColumns]
    print("MAX:")
    print(commonSourceDF["GLdate"].agg('max'))
    print(commonDestinationDF["GLdate"].agg('max'))
    #Keep only new Records
    commonSourceDF['Month'] = pd.to_datetime(commonSourceDF['GLdate'])
    commonDestinationDF['Month'] = pd.to_datetime(commonDestinationDF['GLdate'])
    commonSourceDF = commonSourceDF[commonSourceDF["Month"] > commonDestinationDF["Month"].agg('max')]


    
    if len(commonSourceDF) == 0:
        return pd.DataFrame()
    
    else:
        result = pd.concat([commonDestinationDF,commonSourceDF],axis=0,ignore_index=True)
        
        # Formulas
        for (x,y) in zip(range(0,len(result)),range(1,len(result)+1)):
            result.loc[x,"ProjectName"]= f"=INDEX('Master Data Sheet'!$A:$K,MATCH(Solar!$D{y+1},'Master Data Sheet'!$F:$F,0),1)"
            result.loc[x,"Serial"] = f"=INDEX('Master Data Sheet'!$A:$K,MATCH(Solar!$D{y+1},'Master Data Sheet'!$F:$F,0),8)"
            result.loc[x,"ProjectsManager"] = f"=INDEX('Master Data Sheet'!$A:$K,MATCH(Solar!$D{y+1},'Master Data Sheet'!$F:$F,0),9)"
            result.loc[x,"P/D/O"] = f"=INDEX('Master Data Sheet'!$A:$K,MATCH(Solar!$D{y+1},'Master Data Sheet'!$F:$F,0),11)"
            result.loc[x,"Location "] = f"=INDEX('Master Data Sheet'!$A:$K,MATCH(Solar!$D{y+1},'Master Data Sheet'!$F:$F,0),10)"
            
        
        return result

##############################

#Load
##############################
def SaveDataFrameToExcelSheet(DataFrame,excelFile,sheetName,finalColumns):
    if len(DataFrame) == 0:
        return "No New Data to be added"
    
    else:
        DataFrame = DataFrame[finalColumns]
        container_client = ContainerClient.from_connection_string(conn_str=connect_str,container_name=container_name)
        downloaded_blob = container_client.download_blob(excelFile)
        file = io.BytesIO(downloaded_blob.readall())
        with pd.ExcelWriter(file, engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:
            DataFrame.to_excel(writer,sheet_name=sheetName,index=False)

        
        #Create Excel Table
        tab = openpyxl.worksheet.table.Table(displayName="SolarTable", ref=f'A1:{openpyxl.utils.get_column_letter(DataFrame.shape[1])}{len(DataFrame)+1}')
        wb = load_workbook(file)
        wb[sheetName].add_table(tab)
        wb.save(filename=excelFile)


        blob_client = container_client.get_blob_client(blob=excelFile)
        with open(excelFile,"rb") as data:
            blob_client.upload_blob(data,overwrite=True)
        
        return "DONE <3"

##############################

#Helping Functions
##############################
def remove(s):
    # Using the replace function to remove whitespaces
    s = s.replace(" ", "")
    return s

##############################

sourceExcelFile = "RME_Projects_Cost_Distribution_.xlsb"
destinationExcelFile="Transportation & Equipment 2023.xlsx"
sheetName = "Solar"
source,commonColumn_names = ReadSourceExcelSheetToDataFrame(excelFile=sourceExcelFile)
commonColumns = list(commonColumn_names)
destination,finalColumns = ReadDestinationExcelSheetToDataFrame(excelFile=destinationExcelFile,sheetName=sheetName,columnNames=commonColumn_names)
result = AppendNewData(SourceDF=source,DestinationDF=destination,commonColumns=commonColumns)
print(SaveDataFrameToExcelSheet(DataFrame=result,excelFile=destinationExcelFile,sheetName=sheetName,finalColumns=finalColumns))
