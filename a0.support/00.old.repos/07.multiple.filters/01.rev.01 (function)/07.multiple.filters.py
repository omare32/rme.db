import openpyxl
import os
import glob

def filter_excel_data(words):
    # Get the list of Excel files in the current directory
    excel_files = glob.glob("*.xlsx")

    # Loop through the Excel files
    for excel_file in excel_files:
        # Load the Excel file
        workbook = openpyxl.load_workbook(excel_file)

        # Get the worksheet
        sheet = workbook.active

        # Create a new sheet to store the filtered data
        new_sheet = workbook.create_sheet("Filtered Data")

        # Filter the rows for the specified words
        filtered_rows = []
        for row in sheet.iter_rows():
            if row["Trx Type"].value in words:
                filtered_rows.append(row)

        # Append the filtered rows to the new sheet
        for row in filtered_rows:
            new_sheet.append(row)

        # Save the Excel file
        workbook.save(excel_file)
    
if __name__ == "__main__":
    # Get the words to filter for
    words = ["Move Order Issue on Project", "RME Issue ( On Project)", "RME Site Return"]

    # Filter the Excel data
    filter_excel_data(words)