import openpyxl
import re

# Load the Excel file
wb = openpyxl.load_workbook("03.Kafr Shokr Client Invoice And Cash In Status (2023.04.30).xlsx")

# Get the source sheet
sheet1 = wb.active

# Get the words to find and replace
find_word = "Invoice"
replace_word = "Koko"

# Find and replace the words in all cells
for row in range(1, sheet1.max_row + 1):
    for column in range(1, sheet1.max_column + 1):
        cell_value = sheet1.cell(row=row, column=column).value
        matches = re.findall(find_word, cell_value)
        for match in matches:
            sheet1.cell(row=row, column=column).value = cell_value.replace(match, replace_word)

# Create a new Excel file
wb2 = openpyxl.Workbook()
sheet2 = wb2.active

# Copy the data from the source sheet to the new sheet
for row in range(1, sheet1.max_row + 1):
    for column in range(1, sheet1.max_column + 1):
        sheet2.cell(row=row, column=column).value = sheet1.cell(row=row, column=column).value

# Save the new Excel file
wb2.save("03.Kafr Shokr Client Invoice And Cash In Status (2023.04.30)_New_String.xlsx")