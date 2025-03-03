import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook("03.Kafr Shokr Client Invoice And Cash In Status (2023.04.30).xlsx")

# Get the source sheet
sheet1 = wb.active

# Get the words to find and replace
find_word = "Invoice"
replace_word = "Koko"

# Find and replace the words in all cells
for row in range(1, sheet1.max_row + 1):
    for column in range(1, sheet1.max_column + 1):
        cell_value = sheet1.cell(row=row, column=column).value
        if cell_value == find_word:
            sheet1.cell(row=row, column=column).value = replace_word

# Save the Excel file
wb.save("03.Kafr Shokr Client Invoice And Cash In Status (2023.04.30).new.xlsx")