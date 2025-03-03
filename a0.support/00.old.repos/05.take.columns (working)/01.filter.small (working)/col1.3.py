import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook("small.list.xlsx")

# Get the source sheet
sheet1 = wb.active

# Get the column numbers to copy
column1 = 1
column3 = 3

# Create a new sheet
sheet2 = wb.create_sheet("two.columns")

# Copy the data from the source sheet to the new sheet
for row in range(1, sheet1.max_row + 1):
    sheet2.cell(row=row, column=column1).value = sheet1.cell(row=row, column=column1).value
    sheet2.cell(row=row, column=column3).value = sheet1.cell(row=row, column=column3).value

# Replace column 2 with column 3
for row in range(1, sheet2.max_row + 1):
    sheet2.cell(row=row, column=2).value = sheet2.cell(row=row, column=3).value

# Delete column 3
sheet2.delete_cols(3)

# Save the Excel file
wb.save("small.list.new.xlsx")