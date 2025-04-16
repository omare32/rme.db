import mysql.connector as mysql
import pandas as pd
import matplotlib.pyplot as plt
import arabic_reshaper
from bidi.algorithm import get_display
import os
from datetime import datetime
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter

def fix_arabic_text(text):
    if any('\u0600' <= c <= '\u06FF' for c in str(text)):  # Check if text contains Arabic
        reshaped_text = arabic_reshaper.reshape(str(text))
        bidi_text = get_display(reshaped_text)
        return bidi_text
    return text

def fetch_data(cursor, query, columns):
    cursor.execute(query)
    data = cursor.fetchall()
    df = pd.DataFrame(data, columns=columns)
    df['AMOUNT'] = pd.to_numeric(df['AMOUNT'], errors='coerce')
    return df

def create_bar_plot(data, title, output_path, ylabel):
    plt.figure(figsize=(15, 8))
    
    # Fix Arabic text in labels
    fixed_labels = [fix_arabic_text(label) for label in data.index]
    
    plt.barh(range(len(data)), data.values/1_000_000)
    plt.title(fix_arabic_text(title), fontsize=14)
    plt.xlabel('Amount (Millions EGP)', fontsize=12)
    plt.ylabel(fix_arabic_text(ylabel), fontsize=12)
    plt.yticks(range(len(data)), fixed_labels)
    
    # Add value labels
    for i, v in enumerate(data.values):
        plt.text(v/1_000_000, i, f'{v/1_000_000:,.1f}M', va='center')
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()

def create_comparison_plot(cost_data, cash_data, title, output_path, ylabel):
    plt.figure(figsize=(15, 10))
    
    # Get all unique indices
    all_indices = sorted(set(cost_data.index) | set(cash_data.index))
    
    # Fix Arabic text in labels
    fixed_labels = [fix_arabic_text(label) for label in all_indices]
    
    # Prepare data
    cost_values = [cost_data.get(idx, 0)/1_000_000 for idx in all_indices]
    cash_values = [cash_data.get(idx, 0)/1_000_000 for idx in all_indices]
    
    # Create bars
    y_pos = range(len(all_indices))
    plt.barh([y - 0.2 for y in y_pos], cost_values, 0.4, label='Cost', color='#ff7f0e')
    plt.barh([y + 0.2 for y in y_pos], cash_values, 0.4, label='Cash', color='#1f77b4')
    
    plt.title(fix_arabic_text(title), fontsize=14)
    plt.xlabel('Amount (Millions EGP)', fontsize=12)
    plt.ylabel(fix_arabic_text(ylabel), fontsize=12)
    plt.yticks(y_pos, fixed_labels)
    plt.legend()
    
    # Add value labels
    for i, (cost_v, cash_v) in enumerate(zip(cost_values, cash_values)):
        if cost_v > 0:
            plt.text(cost_v, i - 0.2, f'{cost_v:,.1f}M', va='center')
        if cash_v > 0:
            plt.text(cash_v, i + 0.2, f'{cash_v:,.1f}M', va='center')
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()

def main():
    # Create output directory if it doesn't exist
    output_dir = '03.warehouse/11.cost.vs.cash'
    os.makedirs(output_dir, exist_ok=True)

    # Connection details
    host = "10.10.11.242"
    user = "omar2"
    password = "Omar_54321"
    database = "RME_TEST"

    try:
        print("Connecting to database...")
        connection = mysql.connect(
            host=host,
            user=user,
            password=password,
            database=database
        )
        
        if connection.is_connected():
            print("Connected to MySQL database")

        cursor = connection.cursor()

        # Fetch cost data
        print("Fetching cost data...")
        cost_query = """
        SELECT *
        FROM RME_Projects_Cost_Dist_Line_Report
        WHERE GL_DATE LIKE '2025%'
        AND VENDOR_NAME IS NOT NULL
        AND VENDOR_NAME != ''
        AND VENDOR_NAME NOT IN ('Miscellaneous', 'Petty Cash')
        """
        cursor.execute(cost_query)
        columns = [desc[0] for desc in cursor.description]  # Get all column names
        cost_df_full = pd.DataFrame(cursor.fetchall(), columns=columns)
        
        # Convert GL_DATE to datetime
        cost_df_full['GL_DATE'] = pd.to_datetime(cost_df_full['GL_DATE'])
        
        # Reorder columns to put VENDOR_NAME first
        cols = cost_df_full.columns.tolist()
        cols.remove('VENDOR_NAME')
        cols.remove('GL_DATE')
        new_cols = ['VENDOR_NAME', 'GL_DATE'] + cols
        cost_df_full = cost_df_full[new_cols]
        
        print(f"Fetched {len(cost_df_full)} cost records")

        # Fetch cash data
        print("Fetching cash data...")
        cash_query = """
        SELECT *
        FROM RME_ap_check_payments_Report
        WHERE YEAR(CHECK_DATE) = 2025
        AND `Supplier Name` IS NOT NULL
        AND `Supplier Name` != ''
        """
        cursor.execute(cash_query)
        columns = [desc[0] for desc in cursor.description]  # Get all column names
        cash_df_full = pd.DataFrame(cursor.fetchall(), columns=columns)
        
        # Convert CHECK_DATE to datetime
        cash_df_full['CHECK_DATE'] = pd.to_datetime(cash_df_full['CHECK_DATE'])
        
        # Reorder columns to put Supplier Name first
        cols = cash_df_full.columns.tolist()
        cols.remove('Supplier Name')
        cols.remove('CHECK_DATE')
        new_cols = ['Supplier Name', 'CHECK_DATE'] + cols
        cash_df_full = cash_df_full[new_cols]
        
        print(f"Fetched {len(cash_df_full)} cash records")

        # Export full datasets with proper formatting
        print("\nExporting complete datasets...")
        
        # Export cost data
        cost_output = os.path.join(output_dir, 'complete_cost_data_2025_new.xlsx')
        with pd.ExcelWriter(cost_output, engine='openpyxl') as writer:
            cost_df_full.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            
            # Format date column
            date_col = get_column_letter(cost_df_full.columns.get_loc('GL_DATE') + 1)
            for cell in worksheet[date_col][1:]:  # Skip header
                cell.number_format = 'dd-mmm-yy'
            
            # Format amount columns with thousand separators and no decimals
            amount_columns = ['AMOUNT', 'EQUIV', 'PAYMENT_AMOUNT']
            for col_name in amount_columns:
                if col_name in cost_df_full.columns:
                    col_letter = get_column_letter(cost_df_full.columns.get_loc(col_name) + 1)
                    for cell in worksheet[col_letter][1:]:  # Skip header
                        if cell.value is not None:  # Only format cells with values
                            cell.value = float(cell.value)  # Ensure it's a number
                            cell.number_format = '#,##0'

        # Export cash data
        cash_output = os.path.join(output_dir, 'complete_cash_data_2025_new.xlsx')
        with pd.ExcelWriter(cash_output, engine='openpyxl') as writer:
            cash_df_full.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            
            # Format date column
            date_col = get_column_letter(cash_df_full.columns.get_loc('CHECK_DATE') + 1)
            for cell in worksheet[date_col][1:]:  # Skip header
                cell.number_format = 'dd-mmm-yy'
            
            # Format amount columns with thousand separators and no decimals
            amount_columns = ['AMOUNT', 'EQUIV', 'PAYMENT_AMOUNT']
            for col_name in amount_columns:
                if col_name in cash_df_full.columns:
                    col_letter = get_column_letter(cash_df_full.columns.get_loc(col_name) + 1)
                    for cell in worksheet[col_letter][1:]:  # Skip header
                        if cell.value is not None:  # Only format cells with values
                            cell.value = float(cell.value)  # Ensure it's a number
                            cell.number_format = '#,##0'

        print("Complete datasets exported successfully!")

        # Calculate totals using the correct column names
        cost_project_totals = cost_df_full.groupby('PROJECT_NAME')['AMOUNT'].sum().sort_values(ascending=True)
        cost_vendor_totals = cost_df_full.groupby('VENDOR_NAME')['AMOUNT'].sum().sort_values(ascending=True)
        cash_project_totals = cash_df_full.groupby('PROJECT_NAME')['AMOUNT'].sum().sort_values(ascending=True)
        cash_vendor_totals = cash_df_full.groupby('Supplier Name')['AMOUNT'].sum().sort_values(ascending=True)

        # Get top 10s
        top_10_cost_projects = cost_project_totals.tail(10)
        top_10_cost_vendors = cost_vendor_totals.tail(10)
        top_10_cash_projects = cash_project_totals.tail(10)
        top_10_cash_vendors = cash_vendor_totals.tail(10)

        print("\nCreating visualizations...")
        
        # Create individual plots
        create_bar_plot(
            top_10_cost_projects, 
            'Top 10 Projects by Cost (2025)', 
            os.path.join(output_dir, 'cost_projects_2025.png'),
            'Project Name'
        )
        
        create_bar_plot(
            top_10_cost_vendors, 
            'Top 10 Vendors by Cost (2025)', 
            os.path.join(output_dir, 'cost_vendors_2025.png'),
            'Vendor Name'
        )
        
        create_bar_plot(
            top_10_cash_projects, 
            'Top 10 Projects by Cash (2025)', 
            os.path.join(output_dir, 'cash_projects_2025.png'),
            'Project Name'
        )
        
        create_bar_plot(
            top_10_cash_vendors, 
            'Top 10 Vendors by Cash (2025)', 
            os.path.join(output_dir, 'cash_vendors_2025.png'),
            'Vendor Name'
        )

        # Create comparison plots
        # For projects: combine top 10 from both cost and cash
        top_projects = pd.concat([top_10_cost_projects, top_10_cash_projects])
        top_projects = top_projects.index.unique()
        project_comparison = pd.DataFrame({
            'Cost': cost_project_totals[top_projects],
            'Cash': cash_project_totals[top_projects]
        }).fillna(0)
        
        create_comparison_plot(
            project_comparison['Cost'],
            project_comparison['Cash'],
            'Cost vs Cash by Project (2025)',
            os.path.join(output_dir, 'cost_vs_cash_projects_2025.png'),
            'Project Name'
        )

        # For vendors: combine top 10 from both cost and cash
        top_vendors = pd.concat([top_10_cost_vendors, top_10_cash_vendors])
        top_vendors = top_vendors.index.unique()
        
        # Create a DataFrame with all vendors and fill missing values with 0
        vendor_comparison = pd.DataFrame(index=top_vendors)
        vendor_comparison['Cost'] = cost_vendor_totals.reindex(top_vendors).fillna(0)
        vendor_comparison['Cash'] = cash_vendor_totals.reindex(top_vendors).fillna(0)
        
        create_comparison_plot(
            vendor_comparison['Cost'],
            vendor_comparison['Cash'],
            'Cost vs Cash by Vendor (2025)',
            os.path.join(output_dir, 'cost_vs_cash_vendors_2025.png'),
            'Vendor Name'
        )

        # Print summaries
        print("\nTop 10 Projects by Cost:")
        print(top_10_cost_projects.to_frame().to_string())
        
        print("\nTop 10 Projects by Cash:")
        print(top_10_cash_projects.to_frame().to_string())
        
        print("\nTop 10 Vendors by Cost:")
        print(top_10_cost_vendors.to_frame().to_string())
        
        print("\nTop 10 Vendors by Cash:")
        print(top_10_cash_vendors.to_frame().to_string())

        # Save to Excel
        print("\nSaving detailed data to Excel...")
        
        # Save individual analyses
        with pd.ExcelWriter(os.path.join(output_dir, 'cost_analysis_2025.xlsx')) as writer:
            top_10_cost_projects.to_frame('Amount').to_excel(writer, sheet_name='Top Projects')
            top_10_cost_vendors.to_frame('Amount').to_excel(writer, sheet_name='Top Vendors')
            cost_df_full.to_excel(writer, sheet_name='All Cost Data', index=False)
            
        with pd.ExcelWriter(os.path.join(output_dir, 'cash_analysis_2025.xlsx')) as writer:
            top_10_cash_projects.to_frame('Amount').to_excel(writer, sheet_name='Top Projects')
            top_10_cash_vendors.to_frame('Amount').to_excel(writer, sheet_name='Top Vendors')
            cash_df_full.to_excel(writer, sheet_name='All Cash Data', index=False)
            
        # Save comparison analysis
        with pd.ExcelWriter(os.path.join(output_dir, 'cost_vs_cash_analysis_2025.xlsx')) as writer:
            project_comparison.to_excel(writer, sheet_name='Project Comparison')
            vendor_comparison.to_excel(writer, sheet_name='Vendor Comparison')
        
        print(f"Analysis complete! Results saved in: {output_dir}")

        print("\nGenerating monthly vendor analysis...")
        
        # Create dictionaries to store monthly data
        monthly_cost_data = {}
        monthly_cash_data = {}
        all_vendors = set()
        
        # Process monthly data for each month in 2025
        for month in range(1, 5):  # January to April 2025
            print(f"\nProcessing month {month}/2025...")
            
            # Get cost data for this month
            month_cost_query = """
            SELECT 
                VENDOR_NAME,
                SUM(AMOUNT) as TOTAL_COST
            FROM RME_Projects_Cost_Dist_Line_Report
            WHERE GL_DATE LIKE '2025%'
            AND MONTH(GL_DATE) = %s
            GROUP BY VENDOR_NAME
            """
            cursor.execute(month_cost_query, (month,))
            month_cost_data = pd.DataFrame(cursor.fetchall(), columns=['VENDOR_NAME', 'TOTAL_COST'])
            month_cost_data.set_index('VENDOR_NAME', inplace=True)
            monthly_cost_data[month] = month_cost_data
            all_vendors.update(month_cost_data.index)
            
            # Get cash data for this month
            month_cash_query = """
            SELECT 
                `Supplier Name` as VENDOR_NAME,
                SUM(AMOUNT) as TOTAL_CASH
            FROM RME_ap_check_payments_Report
            WHERE YEAR(CHECK_DATE) = 2025
            AND MONTH(CHECK_DATE) = %s
            GROUP BY `Supplier Name`
            """
            cursor.execute(month_cash_query, (month,))
            month_cash_data = pd.DataFrame(cursor.fetchall(), columns=['VENDOR_NAME', 'TOTAL_CASH'])
            month_cash_data.set_index('VENDOR_NAME', inplace=True)
            monthly_cash_data[month] = month_cash_data
            all_vendors.update(month_cash_data.index)
            
            # Create individual month files as before
            all_month_vendors = pd.Index(set(month_cost_data.index) | set(month_cash_data.index))
            monthly_comparison = pd.DataFrame(index=all_month_vendors)
            monthly_comparison['TOTAL_COST'] = month_cost_data.reindex(all_month_vendors)['TOTAL_COST'].fillna(0)
            monthly_comparison['TOTAL_CASH'] = month_cash_data.reindex(all_month_vendors)['TOTAL_CASH'].fillna(0)
            monthly_comparison['NET_DIFFERENCE'] = monthly_comparison['TOTAL_COST'] - monthly_comparison['TOTAL_CASH']
            
            # Sort by total cost descending
            monthly_comparison.sort_values('TOTAL_COST', ascending=False, inplace=True)
            
            # Export to Excel with formatting
            month_name = ['January', 'February', 'March', 'April'][month-1]
            excel_path = os.path.join(output_dir, f'vendor_analysis_{month_name}_2025.xlsx')
            
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                monthly_comparison.to_excel(writer, sheet_name=f'{month_name} 2025')
                
                # Get the workbook and the worksheet
                workbook = writer.book
                worksheet = writer.sheets[f'{month_name} 2025']
                
                # Format numbers with commas and 2 decimal places
                for col in ['B', 'C', 'D']:  # Columns for TOTAL_COST, TOTAL_CASH, NET_DIFFERENCE
                    for row in range(2, len(monthly_comparison) + 2):
                        cell = worksheet[f'{col}{row}']
                        cell.number_format = '#,##0.00'
                
                # Adjust column widths
                worksheet.column_dimensions['A'].width = 50  # Vendor name column
                worksheet.column_dimensions['B'].width = 20  # Total cost column
                worksheet.column_dimensions['C'].width = 20  # Total cash column
                worksheet.column_dimensions['D'].width = 20  # Net difference column
            
            print(f"Created {excel_path}")
            
            # Print summary
            print(f"\nSummary for {month_name} 2025:")
            print(f"Total vendors: {len(all_month_vendors)}")
            print(f"Total cost: {monthly_comparison['TOTAL_COST'].sum():,.2f}")
            print(f"Total cash: {monthly_comparison['TOTAL_CASH'].sum():,.2f}")

        # Create consolidated report
        print("\nCreating consolidated monthly report...")
        # Remove None values and convert to list for sorting
        valid_vendors = [v for v in all_vendors if v is not None]
        consolidated = pd.DataFrame(index=sorted(valid_vendors))
        
        # Add data for each month
        for month in range(1, 5):
            month_name = ['January', 'February', 'March', 'April'][month-1]
            consolidated[f'Cost_{month_name}'] = monthly_cost_data[month].reindex(consolidated.index)['TOTAL_COST'].fillna(0)
            consolidated[f'Cash_{month_name}'] = monthly_cash_data[month].reindex(consolidated.index)['TOTAL_CASH'].fillna(0)
        
        # Calculate totals
        consolidated['Total_Cost'] = consolidated[[f'Cost_{m}' for m in ['January', 'February', 'March', 'April']]].sum(axis=1)
        consolidated['Total_Cash'] = consolidated[[f'Cash_{m}' for m in ['January', 'February', 'March', 'April']]].sum(axis=1)
        consolidated['Net_Difference'] = consolidated['Total_Cost'] - consolidated['Total_Cash']
        
        # Sort by total cost descending
        consolidated.sort_values('Total_Cost', ascending=False, inplace=True)
        
        # Export consolidated report
        consolidated_path = os.path.join(output_dir, 'vendor_analysis_all_months_2025.xlsx')
        with pd.ExcelWriter(consolidated_path, engine='openpyxl') as writer:
            consolidated.to_excel(writer, sheet_name='All Months 2025')
            
            # Get the workbook and the worksheet
            workbook = writer.book
            worksheet = writer.sheets['All Months 2025']
            
            # Format numbers with commas and 2 decimal places
            for col in range(2, 11):  # All numeric columns
                for row in range(2, len(consolidated) + 2):
                    cell = worksheet.cell(row=row, column=col)
                    cell.number_format = '#,##0.00'
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 50  # Vendor name
            for col in range(2, 11):  # All numeric columns
                worksheet.column_dimensions[chr(64 + col)].width = 20
        
        print(f"Created consolidated report: {consolidated_path}")
        print("\nMonthly vendor analysis complete!")

    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        print(traceback.format_exc())

    finally:
        if 'connection' in locals() and connection.is_connected():
            cursor.close()
            connection.close()
            print("\nMySQL connection is closed")

if __name__ == "__main__":
    main() 