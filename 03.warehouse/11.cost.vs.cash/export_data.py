import mysql.connector as mysql
import pandas as pd
import os
from datetime import datetime
from openpyxl.utils import get_column_letter

def main():
    # Create output directory if it doesn't exist
    output_dir = '03.warehouse/11.cost.vs.cash'
    os.makedirs(output_dir, exist_ok=True)

    # Connection details
    host = "10.10.11.242"
    user = "omar2"
    password = "Omar_54321"
    database = "RME_TEST"

    try:
        print("Connecting to database...")
        connection = mysql.connect(
            host=host,
            user=user,
            password=password,
            database=database
        )
        
        if connection.is_connected():
            print("Connected to MySQL database")

        cursor = connection.cursor()

        # Fetch cost data
        print("Fetching cost data...")
        cost_query = """
        SELECT *
        FROM RME_Projects_Cost_Dist_Line_Report
        WHERE GL_DATE LIKE '2025%'
        AND VENDOR_NAME IS NOT NULL
        AND VENDOR_NAME != ''
        AND VENDOR_NAME NOT IN ('Miscellaneous', 'Petty Cash')
        """
        cursor.execute(cost_query)
        columns = [desc[0] for desc in cursor.description]  # Get all column names
        cost_df_full = pd.DataFrame(cursor.fetchall(), columns=columns)
        
        # Convert GL_DATE to datetime
        cost_df_full['GL_DATE'] = pd.to_datetime(cost_df_full['GL_DATE'])
        
        # Reorder columns to put VENDOR_NAME first
        cols = cost_df_full.columns.tolist()
        cols.remove('VENDOR_NAME')
        cols.remove('GL_DATE')
        new_cols = ['VENDOR_NAME', 'GL_DATE'] + cols
        cost_df_full = cost_df_full[new_cols]
        
        print(f"Fetched {len(cost_df_full)} cost records")

        # Fetch cash data
        print("Fetching cash data...")
        cash_query = """
        SELECT *
        FROM RME_ap_check_payments_Report
        WHERE YEAR(CHECK_DATE) = 2025
        AND `Supplier Name` IS NOT NULL
        AND `Supplier Name` != ''
        """
        cursor.execute(cash_query)
        columns = [desc[0] for desc in cursor.description]  # Get all column names
        cash_df_full = pd.DataFrame(cursor.fetchall(), columns=columns)
        
        # Convert CHECK_DATE to datetime
        cash_df_full['CHECK_DATE'] = pd.to_datetime(cash_df_full['CHECK_DATE'])
        
        # Reorder columns to put Supplier Name first
        cols = cash_df_full.columns.tolist()
        cols.remove('Supplier Name')
        cols.remove('CHECK_DATE')
        new_cols = ['Supplier Name', 'CHECK_DATE'] + cols
        cash_df_full = cash_df_full[new_cols]
        
        print(f"Fetched {len(cash_df_full)} cash records")

        # Export full datasets with proper formatting
        print("\nExporting complete datasets...")
        
        # Export cost data
        cost_output = os.path.join(output_dir, 'complete_cost_data_2025_new.xlsx')
        with pd.ExcelWriter(cost_output, engine='openpyxl') as writer:
            cost_df_full.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            
            # Format date column
            date_col = get_column_letter(cost_df_full.columns.get_loc('GL_DATE') + 1)
            for cell in worksheet[date_col][1:]:  # Skip header
                cell.number_format = 'dd-mmm-yy'
            
            # Format amount columns with thousand separators and no decimals
            amount_columns = ['AMOUNT', 'EQUIV', 'PAYMENT_AMOUNT']
            for col_name in amount_columns:
                if col_name in cost_df_full.columns:
                    col_letter = get_column_letter(cost_df_full.columns.get_loc(col_name) + 1)
                    for cell in worksheet[col_letter][1:]:  # Skip header
                        if cell.value is not None:  # Only format cells with values
                            cell.value = float(cell.value)  # Ensure it's a number
                            cell.number_format = '#,##0'

        # Export cash data
        cash_output = os.path.join(output_dir, 'complete_cash_data_2025_new.xlsx')
        with pd.ExcelWriter(cash_output, engine='openpyxl') as writer:
            cash_df_full.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            
            # Format date column
            date_col = get_column_letter(cash_df_full.columns.get_loc('CHECK_DATE') + 1)
            for cell in worksheet[date_col][1:]:  # Skip header
                cell.number_format = 'dd-mmm-yy'
            
            # Format amount columns with thousand separators and no decimals
            amount_columns = ['AMOUNT', 'EQUIV', 'PAYMENT_AMOUNT']
            for col_name in amount_columns:
                if col_name in cash_df_full.columns:
                    col_letter = get_column_letter(cash_df_full.columns.get_loc(col_name) + 1)
                    for cell in worksheet[col_letter][1:]:  # Skip header
                        if cell.value is not None:  # Only format cells with values
                            cell.value = float(cell.value)  # Ensure it's a number
                            cell.number_format = '#,##0'

        print("Complete datasets exported successfully!")

    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        print(traceback.format_exc())

    finally:
        if 'connection' in locals() and connection.is_connected():
            cursor.close()
            connection.close()
            print("\nMySQL connection is closed")

if __name__ == "__main__":
    main() 